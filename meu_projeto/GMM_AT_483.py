import openpyxl
from datetime import datetime

def ler_dados_origem(caminho_origem):
    """Lê dados de todas as abas da planilha de origem e retorna um dicionário com os dados."""
    wb_origem = openpyxl.load_workbook(caminho_origem)
    dados = {}

    for aba_origem in wb_origem.sheetnames:
        aba = wb_origem[aba_origem]
        dados[aba_origem] = []
        print(f"Planilha de origem: {aba.title}")

        for row in aba.iter_rows(min_row=2, values_only=True):
            print(f"Linhas lidas: {row}")

            # Verifique os índices das colunas aqui. Ajustei para corresponder aos exemplos fornecidos.
            data_str = row[1]  # Supondo que a data esteja na coluna B (índice 1)
            temperatura_str = row[4]  # Supondo que a temperatura esteja na coluna E (índice 4)
            
            print(f"Data extraída: {data_str}")
            print(f"Temperatura extraída: {temperatura_str}")

            try:
                if isinstance(data_str, str):
                    data = datetime.strptime(data_str, '%d/%m/%Y')
                else:
                    print(f"Data inválida encontrada e ignorada: {data_str}")
                    continue
            except ValueError:
                print(f"Data inválida encontrada e ignorada: {data_str}")
                continue
            
            try:
                if isinstance(temperatura_str, str):
                    temperatura_atual = float(temperatura_str.replace(' °C', '').replace(' ºC', '').replace(',', '.'))
                else:
                    print(f"Temperatura inválida encontrada e ignorada: {temperatura_str}")
                    continue
            except ValueError:
                print(f"Temperatura inválida encontrada e ignorada: {temperatura_str}")
                continue
            
            dados[aba_origem].append((data, temperatura_atual))
    
    print("Dados lidos:", dados)
    return dados

def encontrar_linha_vazia_tabela(aba):
    """Encontra a primeira linha vazia na coluna B dentro da tabela."""
    for row in range(17, aba.max_row + 1):
        if aba[f'B{row}'].value is None:
            return row
    return aba.max_row + 1

def linha_existe(aba, data, temperatura):
    """Verifica se a combinação de data e temperatura já existe na aba."""
    for row in range(17, aba.max_row + 1):
        data_existente = aba[f'B{row}'].value
        temperatura_existente = aba[f'C{row}'].value
        
        if data_existente == data.strftime('%d/%m/%Y') and temperatura_existente == temperatura:
            return True
    return False

def encontrar_aba_destino(wb_destino, aba_nome_origem):
    """Encontra a aba na planilha de destino com base no prefixo da aba de origem."""
    prefixo = aba_nome_origem.split(' ')[0]  # Corrigido para o prefixo correto
    for aba_nome_destino in wb_destino.sheetnames:
        if prefixo in aba_nome_destino:
            return wb_destino[aba_nome_destino]
    return None

def transferir_dados(caminho_destino, dados):
    """Preenche os dados na planilha de destino nas abas corretas e linhas corretas."""
    wb_destino = openpyxl.load_workbook(caminho_destino)

    for aba_nome_origem, dados_aba in dados.items():
        aba_destino = encontrar_aba_destino(wb_destino, aba_nome_origem)
        if aba_destino:
            print(f"Aba ativa de destino: {aba_destino.title}")

            for data, temperatura_atual in dados_aba:
                if not linha_existe(aba_destino, data, temperatura_atual):
                    linha = encontrar_linha_vazia_tabela(aba_destino)
                    aba_destino[f'B{linha}'] = data.strftime('%d/%m/%Y')
                    aba_destino[f'C{linha}'] = temperatura_atual
                    print(f"Dados escritos na linha {linha}: Data = {data.strftime('%d/%m/%Y')}, Temperatura = {temperatura_atual}")
                else:
                    print(f"Dados já existentes e ignorados: Data = {data.strftime('%d/%m/%Y')}, Temperatura = {temperatura_atual}")
        else:
            print(f"Aba {aba_nome_origem} não encontrada na planilha de destino.")

    wb_destino.save(caminho_destino)
    print("Planilha salva com sucesso!")

# Caminhos para as planilhas
caminho_origem = r'C:\Users\victo\OneDrive\Documentos\Controle de estoque\projeto_gestao_estoque\dados.xlsx'
caminho_destino = r'C:\Users\victo\OneDrive\Área de Trabalho\GMM-AT-483\MM-05-AT-483 - Carta Controle de Exatidão Equipamentos de Laboratório.xlsx'

# Ler dados da planilha de origem
dados = ler_dados_origem(caminho_origem)
print("Dados lidos após chamada da função:", dados)

# Transferir os dados para a planilha de destino
transferir_dados(caminho_destino, dados)
