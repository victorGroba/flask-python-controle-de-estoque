from flask import Flask, render_template, request, redirect, url_for
import pandas as pd
from openpyxl import load_workbook, Workbook
import os
from datetime import datetime

app = Flask(__name__)

# Caminho do arquivo Excel
EXCEL_FILE_PATH = r'C:\Users\victo\OneDrive\Documentos\Controle de estoque\projeto_gestao_estoque\dados.xlsx'

# Função para traduzir o nome dos meses para português
def traduzir_mes(mes_ingles):
    meses_pt_br = {
        'January': 'Janeiro', 'February': 'Fevereiro', 'March': 'Março',
        'April': 'Abril', 'May': 'Maio', 'June': 'Junho',
        'July': 'Julho', 'August': 'Agosto', 'September': 'Setembro',
        'October': 'Outubro', 'November': 'Novembro', 'December': 'Dezembro'
    }
    return meses_pt_br.get(mes_ingles, mes_ingles)

@app.route('/')
def index():
    return render_template('form1.html')

@app.route('/registrar', methods=['POST'])
def registrar():
    data = {
        "Equipamento": request.form['equipment'],
        "Data": datetime.strptime(request.form['date'], '%Y-%m-%d').strftime('%d/%m/%Y'),
        "Hora": request.form['time'],
        "Responsável": request.form['responsible'],
        "Temperatura Atual": request.form.get('current-temp', 'N/A'),
        "Temperatura Máxima": request.form.get('max-temp', 'N/A'),
        "Temperatura Mínima": request.form.get('min-temp', 'N/A'),
        "Observações": request.form.get('observations', 'Nenhum'),
        "Observações Texto": request.form.get('observations-text', '')
    }
    
    sheet_name = request.form['equipment'].strip()  # Remove espaços em branco adicionais

    try:
        # Cria um novo arquivo se não existir
        if not os.path.exists(EXCEL_FILE_PATH):
            book = Workbook()
            book.save(EXCEL_FILE_PATH)
        
        # Carrega o arquivo existente
        book = load_workbook(EXCEL_FILE_PATH)
        
        # Se a aba não existir, cria uma nova
        if sheet_name not in book.sheetnames:
            sheet = book.create_sheet(title=sheet_name)
            # Adiciona cabeçalhos na nova aba
            headers = list(data.keys())
            for col_num, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col_num, value=header)
        else:
            sheet = book[sheet_name]
        
        # Encontra a próxima linha vazia
        next_row = sheet.max_row + 1
        
        # Adiciona o novo registro na próxima linha
        for col_num, header in enumerate(data.keys(), 1):
            sheet.cell(row=next_row, column=col_num, value=data[header])
        
        # Salva as mudanças
        book.save(EXCEL_FILE_PATH)

    except Exception as e:
        print(f"Erro ao manipular o arquivo Excel: {e}")
        return f"Ocorreu um erro ao processar o arquivo: {e}", 500

    return redirect(url_for('index'))

@app.route('/relatorio')
def relatorio():
    try:
        # Verifica se o arquivo existe
        if not os.path.exists(EXCEL_FILE_PATH):
            return "O arquivo de dados não existe.", 404

        # Carrega todos os dados das planilhas
        df = pd.read_excel(EXCEL_FILE_PATH, engine='openpyxl', sheet_name=None)
        instrument_reports = []

        for sheet_name, data in df.items():
            if 'Data' not in data.columns:
                print(f"Coluna 'Data' não encontrada na aba '{sheet_name}'.")
                continue

            # Converte a data para o formato datetime
            try:
                data['Data'] = pd.to_datetime(data['Data'], format='%d/%m/%Y', errors='coerce')
                data = data.dropna(subset=['Data'])  # Remove entradas com data inválida
                # Formata a data para exibição
                data['Data'] = data['Data'].dt.strftime('%d/%m/%Y')
            except Exception as e:
                print(f"Erro ao formatar a data na aba '{sheet_name}': {e}")
                continue

            # Ordena os dados por data
            data = data.sort_values(by='Data')

            # Formata temperaturas
            for col in ['Temperatura Atual', 'Temperatura Máxima', 'Temperatura Mínima']:
                if col in data.columns:
                    # Converte para numérico e formata
                    data[col] = pd.to_numeric(data[col], errors='coerce')
                    data[col] = data[col].apply(lambda x: f"{x:.1f} ºC" if pd.notnull(x) else "N/A")

            # Adiciona o relatório à lista de relatórios
            for month_year, group in data.groupby(data['Data'].str[3:]):  # Agrupa por mês e ano
                mes_ingles = datetime.strptime(month_year, '%m/%Y').strftime('%B')
                mes_portugues = traduzir_mes(mes_ingles)
                report_name = f"Termômetro - {sheet_name}"  # Nome do equipamento com prefixo
                instrument_reports.append({
                    'sheet_name': report_name,
                    'month_year': f"{mes_portugues} {datetime.strptime(month_year, '%m/%Y').year}",
                    'html_table': group.to_html(classes='table table-striped table-bordered text-center', index=False, header=True),
                    'month_year_full': f"{mes_portugues}/{datetime.strptime(month_year, '%m/%Y').year}"
                })

        if not instrument_reports:
            return "Nenhum dado encontrado para exibição.", 404

        return render_template('relatorio_temperatura.html', instrument_reports=instrument_reports)
    except Exception as e:
        print(f"Erro ao ler o arquivo Excel para o relatório: {e}")
        return f"Ocorreu um erro ao processar o relatório: {e}", 500

@app.route('/relatorio_individual/<sheet_name>/<month_year>')
def relatorio_individual(sheet_name, month_year):
    try:
        if not os.path.exists(EXCEL_FILE_PATH):
            return "O arquivo de dados não existe.", 404

        # Remove espaços em branco adicionais do nome da aba
        sheet_name = sheet_name.strip()

        # Carrega o arquivo Excel
        df = pd.read_excel(EXCEL_FILE_PATH, engine='openpyxl', sheet_name=None)

        # Verifica se a aba existe
        matching_sheets = [name for name in df.keys() if sheet_name in name]
        if not matching_sheets:
            return f"Aba '{sheet_name}' não encontrada no arquivo.", 404

        # Carrega a aba específica
        df_sheet = df[matching_sheets[0]]

        if 'Data' not in df_sheet.columns:
            return "Coluna 'Data' não encontrada.", 404

        # Converte e formata a coluna 'Data'
        try:
            df_sheet['Data'] = pd.to_datetime(df_sheet['Data'], format='%d/%m/%Y', errors='coerce')
            df_sheet = df_sheet.dropna(subset=['Data'])
            df_sheet['Data'] = df_sheet['Data'].dt.strftime('%d/%m/%Y')
        except Exception as e:
            return f"Erro ao formatar a data: {e}", 500

        # Verifica o formato de month_year
        try:
            month_year_datetime = datetime.strptime(month_year, '%m/%Y')
            month_year_formatted = month_year_datetime.strftime('%m/%Y')
            print(f"month_year: {month_year}, month_year_formatted: {month_year_formatted}")
        except ValueError:
            return "Formato de mês e ano inválido. Use o formato MM/YYYY.", 400

        # Filtra os dados para o mês e ano especificados
        df_sheet_filtered = df_sheet[df_sheet['Data'].str[3:] == month_year_formatted]

        if df_sheet_filtered.empty:
            print(f"Nenhum dado encontrado para o período especificado. month_year: {month_year_formatted}")
            return "Nenhum dado encontrado para o período especificado.", 404
        
        df_sheet_filtered = df_sheet_filtered.sort_values(by='Data')
        
        for col in ['Temperatura Atual', 'Temperatura Máxima', 'Temperatura Mínima']:
            if col in df_sheet_filtered.columns:
                df_sheet_filtered[col] = pd.to_numeric(df_sheet_filtered[col], errors='coerce')
                df_sheet_filtered[col] = df_sheet_filtered[col].apply(lambda x: f"{x:.1f} ºC" if pd.notnull(x) else "N/A")
        
        html_table = df_sheet_filtered.to_html(classes='table table-striped table-bordered text-center', index=False, header=True)
        
        return render_template('relatorio_individual.html', 
                               sheet_name=sheet_name, 
                               month_year=month_year, 
                               html_table=html_table)
    except Exception as e:
        print(f"Erro ao processar o relatório: {e}")
        return f"Ocorreu um erro ao processar o relatório: {e}", 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5001, debug=True)
