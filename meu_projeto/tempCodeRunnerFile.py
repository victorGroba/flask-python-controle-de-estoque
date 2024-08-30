from flask import Flask, render_template, request, redirect, url_for
import pandas as pd
from openpyxl import load_workbook, Workbook
import os
from datetime import datetime

app = Flask(__name__)

# Caminho do arquivo Excel
EXCEL_FILE_PATH = r'C:\Users\victo\OneDrive\Documentos\Controle de estoque\projeto_gestao_estoque\dados.xlsx'

@app.route('/')
def index():
    return render_template('form1.html')

@app.route('/registrar', methods=['POST'])
def registrar():
    data = {
        "Equipamento": request.form['equipment'],
        "Especificação": request.form['specification'],
        "Erro Padrão (°C)": request.form['default-error'],
        "Data": datetime.strptime(request.form['date'], '%Y-%m-%d').strftime('%d/%m/%Y'),
        "Hora": request.form['time'],
        "Responsável": request.form['responsible'],
        "Temperatura Atual": f"{float(request.form['current-temp']):.1f} ºC",
        "Temperatura Máxima": f"{float(request.form['max-temp']):.1f} ºC",
        "Temperatura Mínima": f"{float(request.form['min-temp']):.1f} ºC",
        "Observações": request.form['observations']
    }
    
    # Nome da aba e sufixo
    equipment_name = request.form['equipment']
    sheet_name = f"GMM{equipment_name.replace(' ', '')}"
    
    # Data para criação de nome de planilha mensal
    report_date = datetime.strptime(request.form['date'], '%Y-%m-%d')
    month_year = report_date.strftime('%Y-%m')  # Formato YYYY-MM para agrupamento mensal

    try:
        # Cria um novo arquivo se não existir
        if not os.path.exists(EXCEL_FILE_PATH):
            book = Workbook()
            book.save(EXCEL_FILE_PATH)
        
        # Carrega o arquivo existente
        book = load_workbook(EXCEL_FILE_PATH)
        
        # Cria a aba para o mês se não existir
        monthly_sheet_name = f"{month_year} - {sheet_name}"
        if monthly_sheet_name not in book.sheetnames:
            sheet = book.create_sheet(title=monthly_sheet_name)
            # Adiciona cabeçalhos na nova aba
            headers = list(data.keys())
            for col_num, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col_num, value=header)
        else:
            sheet = book[monthly_sheet_name]
        
        # Encontra a próxima linha vazia
        next_row = sheet.max_row + 1
        
        # Adiciona o novo registro na próxima linha
        for col_num, header in enumerate(data.keys(), 1):
            sheet.cell(row=next_row, column=col_num, value=data[header])
        
        # Salva as mudanças
        book.save(EXCEL_FILE_PATH)

    except Exception as e:
        print(f"Erro ao manipular o arquivo Excel: {e}")
        return f"Ocorreu um erro ao processar o arquivo: {e}", 500

    return redirect(url_for('index'))

@app.route('/relatorio')
def relatorio():
    try:
        # Verifica se o arquivo existe
        if not os.path.exists(EXCEL_FILE_PATH):
            return "O arquivo de dados não existe.", 404

        # Carrega todos os dados das planilhas
        df = pd.read_excel(EXCEL_FILE_PATH, engine='openpyxl', sheet_name=None)
        tables = []
        for sheet_name, data in df.items():
            # Verifica se a coluna 'Data' existe
            if 'Data' not in data.columns:
                print(f"Coluna 'Data' não encontrada na aba '{sheet_name}'.")
                continue

            # Converte a data do formato original para o formato desejado
            try:
                data['Data'] = pd.to_datetime(data['Data'], format='%d/%m/%Y', errors='coerce').dt.strftime('%d/%m/%Y')
            except Exception as e:
                print(f"Erro ao formatar a data na aba '{sheet_name}': {e}")
                continue
            
            # Formata temperaturas
            if 'Temperatura Atual' in data.columns:
                data['Temperatura Atual'] = data['Temperatura Atual'].apply(lambda x: f"{x:.1f} ºC" if pd.notnull(x) else "N/A")
            if 'Temperatura Máxima' in data.columns:
                data['Temperatura Máxima'] = data['Temperatura Máxima'].apply(lambda x: f"{x:.1f} ºC" if pd.notnull(x) else "N/A")
            if 'Temperatura Mínima' in data.columns:
                data['Temperatura Mínima'] = data['Temperatura Mínima'].apply(lambda x: f"{x:.1f} ºC" if pd.notnull(x) else "N/A")
            
            # Adiciona a tabela à lista de tabelas
            # Sufixo dos 4 últimos caracteres do nome do equipamento
            suffix = sheet_name[-4:]
            report_title = f"{sheet_name} - {suffix}"
            tables.append({
                'sheet_name': report_title,
                'html_table': data.to_html(classes='table table-striped table-bordered text-center', index=False, header=True)
            })

        if not tables:
            return "Nenhum dado encontrado para exibição.", 404

        return render_template('relatorio_temperatura.html', tables=tables)
    except Exception as e:
        print(f"Erro ao ler o arquivo Excel para o relatório: {e}")
        return f"Ocorreu um erro ao processar o relatório: {e}", 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5001, debug=True)
